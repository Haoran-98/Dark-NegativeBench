import json
import logging
import math
import os
import time

import openpyxl
from openai import OpenAI
import httpx
import requests
import google.generativeai as genai
from openpyxl.styles import Font
from tencentcloud.common import credential
from tencentcloud.common.profile.client_profile import ClientProfile
from tencentcloud.common.profile.http_profile import HttpProfile
from tencentcloud.hunyuan.v20230901 import hunyuan_client, models
from dashscope import Generation
import dashscope

QUESTIONNAIRES_ADDREESS = "data/questionnaires"


def log(psychology_scale_name, model_name, language):
    """
    Log
    @return:log object
    """
    logger = logging.getLogger('my_logger')
    logger.setLevel(logging.DEBUG)
    
    file_handler = logging.FileHandler(f"log/{psychology_scale_name}_{model_name}_{language}_chat.log")
    console_handler = logging.StreamHandler()
    
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)
    
    logger.addHandler(file_handler)
    
    return logger


def get_psychology_scale_object_by_name(psychology_scale_name):
    """
    Get scale information in the form of a json object by scale name
    @param psychology_scale_name:psychology scale name
    @return:psychology scale json object(Dictionary object)
    """
    with open(
            f'{QUESTIONNAIRES_ADDREESS}/{psychology_scale_name}.json', 'r',
            encoding='utf-8') as psychology_scale_json_file:
        psychology_scale_object = json.load(psychology_scale_json_file)
        return psychology_scale_object


def get_psychology_scale_prompt(psychology_scale_object, language):
    """
    Get psychology scale prompt from the psychology scale json object
    @param psychology_scale_object: psychology scale json object
    @param language:language of the psychology scale(en or zh)
    @return: prompt(str)
    """
    if language == 'en':
        return psychology_scale_object['psychobench_prompt_en']
    elif language == 'zh':
        return psychology_scale_object['psychobench_prompt_zh']
    else:
        print('Invalid language')
        return None


def get_psychology_scale_each_sentence_str_list(psychology_scale_name, language):
    """
    get psychology scale each sentence str
    @param psychology_scale_name:
    @param language:
    @return: scale each sentence str(list)
    """
    psychology_scale_object = get_psychology_scale_object_by_name(psychology_scale_name)
    questions_dict = psychology_scale_object['questions']
    questions_list_en_str_list = []
    questions_list_zh_str_list = []
    for i in range(1, len(questions_dict) + 1):
        questions_list_en_str_list.append(questions_dict[str(i)]['origin_en'])
        questions_list_zh_str_list.append(questions_dict[str(i)]['origin_zh'])
    if language == 'zh':
        return questions_list_zh_str_list
    else:
        return questions_list_en_str_list


def get_api_base_by_model_name(model_name):
    """
    get API_BASE url
    @param model_name:
    @return: API_BASE url
    """
    BaiDu_AI_URL = f"https://aip.baidubce.com/rpc/2.0/ai_custom/v1/wenxinworkshop/chat/{model_name}?access_token="
    switch_dict = {
        'glm-3-turbo': 'https://open.bigmodel.cn/api/paas/v4',
        'glm-4': 'https://open.bigmodel.cn/api/paas/v4',
        'moonshot-v1-8k': 'https://api.moonshot.cn/v1',
        'Baichuan2-Turbo': 'https://api.baichuan-ai.com/v1/chat/completions',
        'llama_2_7b': BaiDu_AI_URL,
        'llama_2_13b': BaiDu_AI_URL,
        'llama_2_70b': BaiDu_AI_URL,
        'llama_3_8b': BaiDu_AI_URL,
        'llama_3_70b': BaiDu_AI_URL,
        'ernie-3.5-8k-0205': BaiDu_AI_URL,
        'hunyuan-pro': 'https://hunyuan.tencentcloudapi.com'
    }
    return switch_dict.get(model_name)


def llms_api(psychology_scale_name, model_name, key, language):
    """
    Calling the LLMs API for evaluation
    @param psychology_scale_name:
    @param language:
    @param model_name:
    @param key:
    @return:
    """
    logger = log(psychology_scale_name, model_name, language)
    if language == 'zh':
        system_content = get_common_prompts_json()['json_data_prefix_zh_for_system']
        json_data_suffix = get_common_prompts_json()['json_data_suffix_zh']
    else:
        system_content = get_common_prompts_json()['json_data_prefix_en_for_system']
        json_data_suffix = get_common_prompts_json()['json_data_suffix_en']
    psychology_scale_json = json.dumps(get_psychology_scale_each_sentence_str_list(psychology_scale_name, language))
    prompt = get_psychology_scale_prompt(get_psychology_scale_object_by_name(psychology_scale_name), language)
    api_url = get_api_base_by_model_name(model_name)
    response_content = ""
    if model_name == 'gpt-3.5-turbo' or model_name == 'moonshot-v1-8k' or model_name == 'glm-3-turbo' or model_name == 'glm-4':
        # when model_name == 'gpt-3.5-turbo' api_url = None, base_url = None, Call the official API of Open
        client = OpenAI(
            base_url=api_url,
            api_key=key,
            http_client=httpx.Client(
                follow_redirects=True,
            ),
        )
        history = [
            {"role": "system", "content": system_content},
            {"role": "user",
             "content": prompt},
        ]
        response = client.chat.completions.create(
            model=model_name,
            messages=history,
            temperature=0.01,
        )
        history.append({"role": "assistant", "content": response.choices[0].message.content})
        history.append(
            {"role": "user",
             "content": f"{json.dumps(psychology_scale_json)}{json_data_suffix}"})
        response = client.chat.completions.create(
            model=model_name,
            messages=history,
            temperature=0.01,
        )
        response_content = response.choices[0].message.content
        logger.info("response_content: " + response_content)
        history.append({"role": "assistant", "content": response_content})
        logger.info("history: " + json.dumps(history))
    elif model_name == "Baichuan2-Turbo":
        history = [
            {
                "role": "user",
                "content": prompt
            }
        ]
        headers = {
            "Content-Type": "application/json",
            "Authorization": "Bearer " + key
        }
        payload = json.dumps({
            "model": "Baichuan2-Turbo",
            "messages": history,
            "temperature": 0.01,
        })
        response = requests.request("POST", api_url, headers=headers, data=payload)
        history.append(
            {"role": "assistant", "content": response.json().get("choices")[0].get("message").get("content")})
        history.append(
            {"role": "user",
             "content": f"{json.dumps(psychology_scale_json)}{json_data_suffix}"})
        payload = json.dumps({
            "model": "Baichuan2-Turbo",
            "messages": history,
            "temperature": 0.01,
        })
        response = requests.request("POST", get_api_base_by_model_name(model_name), headers=headers, data=payload)
        response_content = response.json().get("choices")[0].get("message").get("content")
        logger.info("response_content: " + response_content)
        history.append(
            {"role": "assistant", "content": response_content})
        logger.info("history: " + json.dumps(history))
    elif model_name == "llama_2_7b" or model_name == "llama_2_13b" or model_name == "llama_2_70b" or model_name == "llama_3_8b" or model_name == "llama_3_70b" or model_name == "ernie-3.5-8k-0205":
        history = [
            {
                "role": "user",
                "content": prompt
            }
        ]
        headers = {
            'Content-Type': 'application/json'
        }
        payload = json.dumps({
            "messages": history,
            "temperature": 0.01,
        })
        response = requests.request("POST", api_url + get_access_token(key), headers=headers, data=payload)
        history.append({"role": "assistant", "content": response.json().get("results")})
        psychology_scale_json = psychology_scale_json
        history.append(
            {"role": "user",
             "content": f"{json.dumps(psychology_scale_json)}{json_data_suffix}"})
        payload = json.dumps({
            "messages": history,
            "temperature": 0.01,
        })
        response = requests.request("POST", get_api_base_by_model_name(model_name) + get_access_token(key),
                                    headers=headers, data=payload)
        response_content = response.json().get("results")
        logger.info("response_content: " + response_content)
        history.append({"role": "assistant", "content": response_content})
        logger.info("history: " + json.dumps(history))
    elif model_name == "gemini-1.0-pro" or model_name == "gemini-1.5-pro":
        genai.configure(api_key=api_url, transport="rest")
        # Set up the model
        generation_config = {
            "temperature": 0.01,
            "max_output_tokens": 2048,
        }
        safety_settings = [
            {
                "category": "HARM_CATEGORY_HARASSMENT",
                "threshold": "BLOCK_NONE"
            },
            {
                "category": "HARM_CATEGORY_HATE_SPEECH",
                "threshold": "BLOCK_NONE"
            },
            {
                "category": "HARM_CATEGORY_SEXUALLY_EXPLICIT",
                "threshold": "BLOCK_NONE"
            },
            {
                "category": "HARM_CATEGORY_DANGEROUS_CONTENT",
                "threshold": "BLOCK_NONE"
            }]
        model = genai.GenerativeModel(model_name=model_name,
                                      generation_config=generation_config,
                                      safety_settings=safety_settings
                                      )
        history = [
            {
                "role": "user",
                "content": prompt
            }
        ]
        chat = model.start_chat(history=[])
        chat.send_message(prompt)
        for message in chat.history:
            logger.info(f'**{message.role}**: {message.parts[0].text}')
        chat.send_message(f"{json.dumps(psychology_scale_json)}{json_data_suffix}")
        for message in chat.history:
            logger.info(f'**{message.role}**: {message.parts[0].text}')
            response_content = message.parts[0].text
        logger.info("history: " + json.dumps(history))
    elif model_name == "hunyuan_pro":
        if key.find(',') != -1 and len(key.split(',')) == 2:
            secret_id = key.split(',')[0]
            secret_key = key.split(',')[1]
            cred = credential.Credential(secret_id, secret_key)
        else:
            logger.error(f"Invalid key {key}")
        httpProfile = HttpProfile()
        httpProfile.endpoint = api_url
        clientProfile = ClientProfile()
        clientProfile.httpProfile = httpProfile
        client = hunyuan_client.HunyuanClient(cred, "ap-beijing", clientProfile)
        req = models.ChatCompletionsRequest()
        history = [
            {"Role": "system", "Content": system_content},
            {"Role": "user", "Content": prompt},
        ]
        
        params = {
            "Model": model_name,
            "Temperature": 0.01,
            "Messages": history
        }
        req.from_json_string(json.dumps(params))
        resp = client.ChatCompletions(req)
        resp_json = json.loads(resp.to_json_string())
        history.append({"Role": "assistant", "Content": resp_json.get("Choices")[0].get("Message").get("Content")})
        history.append(
            {"Role": "user",
             "Content": f"{json.dumps(psychology_scale_json)}{json_data_suffix}"})
        req.from_json_string(json.dumps(params))
        resp = client.ChatCompletions(req)
        resp_json = json.loads(resp.to_json_string())
        logger.info(resp_json.get("Choices")[0].get("Message").get("Content"))
        history.append({"Role": "assistant", "Content": resp_json.get("Choices")[0].get("Message").get("Content")})
        logger.info("history: " + json.dumps(history))
        response_content = resp_json.get("Choices")[0].get("Message").get("Content")
    elif model_name == 'qwen-max':
        dashscope.api_key = key
        history = [
            {"role": "system", "content": system_content},
            {"role": "user", "content": prompt},
        ]
        response = Generation.call(model="qwen-max",
                                   messages=history,
                                   temperature=0.01,
                                   # 将输出设置为"message"格式
                                   result_format='message')
        history.append({"role": "assistant", "content": response.output.choices[0]['message']['content']})
        history.append(
            {"role": "user",
             "content": f"{json.dumps(psychology_scale_json)}{json_data_suffix}"})
        response = Generation.call(model="qwen-max",
                                   messages=history,
                                   temperature=0.01,
                                   result_format='message')
        history.append({"role": "assistant", "content": response.output.choices[0]['message']['content']})
        logger.info(response.output.choices[0]['message']['content'])
        logger.info("history: " + json.dumps(history))
        response_content = response.output.choices[0]['message']['content']
    # This can be flexibly adjusted according to the results of different models
    if response_content.find('[') != -1:
        # when response is :
        """
        [1, 1, 2, 1, 2, 1, 7, 1, 3, 5, 5, 5, 4, 4, 3, 4, 2, 6, 2, 2, 2, 1, 3, 3, 2, 1, 1, 1, 2]
        ......
        """
        response_content = response_content.replace('\n', '').replace('，', ',').replace(
            '```json', '').replace('```', '').replace(' ', '')
        # when response is :
        """
        ```json
        [
            {"key": "Once in a while I can't control the urge to strike another person.", "value": 1},
            {"key": "Given enough provocation, I may hit another person.", "value": 1},
            {"key": "If somebody hits me, I hit back.", "value": 2}
            ......
        ]
        ```
        """
        if response_content.find('key') != -1 and response_content.find('value') != -1:
            response_content_object = json.loads(response_content)
            response_content_score_list = []
            for sentence in response_content_object:
                response_content_score_list.append(int(sentence['value']))
            response_content = json.dumps(response_content_score_list)
    else:
        logger.warning(
            "Due to the unstable and uncertain response results of LLMs, this issue has arisen. "
            "There is no list in response_content, you may need to combine it with the log file to get the "
            "correct results" + "response_content: " + response_content)
    # If json. decoder is displayed JSONDecodeError can be checked by examining the relevant logs, as the content of
    # the model's answers is uncertain, it cannot be guaranteed that each answer will be reasonable and effective
    response_content_object = json.loads(response_content)
    logger.info(response_content)
    time.sleep(10)
    return response_content_object


def save_result_json_file(response_content_list, questionnaires_name, model_name, language, run_count):
    """
    Save the evaluation results of a certain scale by the model to "results/model_name/questionnaires_name/result.json"
    @param response_content_list:
    @param run_count:
    @param language:
    @param questionnaires_name:
    @param model_name:
    @return:
    """
    result_json_data = {'scale_name': questionnaires_name, 'model_name': model_name, 'scores': response_content_list}
    if not os.path.exists(f"results/{model_name}/{questionnaires_name}"):
        os.makedirs(f"results/{model_name}/{questionnaires_name}")
    with open(f"results/{model_name}/{questionnaires_name}/result_{language}_{run_count}.json", 'w',
              encoding='utf-8') as result_json_file:
        json.dump(result_json_data, result_json_file)


def calculate_result_json_file_score(questionnaires_name, model_name, language, run_count):
    """
    Calculate the average score and score variance of the model on the scale, and save the results to results/
    model_name/questionnaires_name/score_result_language.json
    @param run_count:
    @param questionnaires_name:
    @param model_name:
    @param language:
    @return:
    """
    score_result = {}
    with open(
            f"results/{model_name}/{questionnaires_name}/result_{language}_{run_count}.json", 'r',
            encoding='utf-8') as f:
        result_data = json.load(f)
    scores_num_list = result_data['scores']
    # Calculate the average score
    Average_score_list = [sum(item) / len(scores_num_list) for item in zip(*scores_num_list)]
    questionnaires_questions = get_psychology_scale_each_sentence_str_list(questionnaires_name, language)
    questions = []
    for i in range(len(questionnaires_questions)):
        question = {'number': str(i + 1), 'question': questionnaires_questions[i], 'score': Average_score_list[i]}
        questions.append(question)
    # Dealing with flipping issues
    scale_object = get_psychology_scale_object_by_name(questionnaires_name)
    reverse = scale_object['reverse']
    if len(reverse) != 0:
        for question in questions:
            if int(question['number']) in reverse:
                question['score'] = float(scale_object['scale']) - float(question['score'])
        for single_score_num in scores_num_list:
            for index, value in enumerate(single_score_num):
                if index + 1 in reverse:
                    single_score_num[index] = float(scale_object['scale']) - single_score_num[index]
    # Calculate the sub scale score and total score for each sample
    score_result['categories'] = []
    categories = scale_object['categories']
    if len(categories) == 0:
        sum_question = 0.0
        for item in questions:
            sum_question = item['score'] + sum_question
        mean = sum_question / len(questions)
        single_score_list = []
        for single_score_num in scores_num_list:
            score = 0.0
            for index, value in enumerate(single_score_num):
                score = score + float(value)
            single_score_list.append(score)
        var = 0.0
        for score in single_score_list:
            if questionnaires_name in ['DTDD', 'Empathy']:
                var = var + (score / len(questions) - mean) * (score / len(questions) - mean)
            else:
                var = var + (score - sum_question) * (score - sum_question)
        sd = math.sqrt(var)
        score_result['categories'].append({"score": {'cat_question_num': len(questions),
                                                     'cat_score_total': sum_question,
                                                     'cat_score_avg': mean,
                                                     'cat_score_sd': sd}})
    else:
        for category in categories:
            cat_name = category['cat_name']
            cat_questions = category['cat_questions']
            cat_score_total = 0
            for question in questions:
                if int(question['number']) in cat_questions:
                    cat_score_total = cat_score_total + float(question['score'])
            mean = cat_score_total / len(cat_questions)
            single_category_score_list = []
            for single_score_num in scores_num_list:
                single_category_score_total = 0.0
                for index, value in enumerate(single_score_num):
                    if index + 1 in cat_questions:
                        single_category_score_total = single_category_score_total + float(value)
                single_score_num_mean = single_category_score_total / len(cat_questions)
                single_category_score = {
                    'cat_score_total': single_category_score_total,
                    'cat_score_avg': single_score_num_mean
                }
                single_category_score_list.append(single_category_score)
            var = 0.0
            for score_item in single_category_score_list:
                if questionnaires_name in ['DTDD', 'Empathy']:
                    var = var + (score_item['cat_score_avg'] - mean) * (score_item['cat_score_avg'] - mean)
                else:
                    var = var + (score_item['cat_score_total'] - cat_score_total) * (
                            score_item['cat_score_total'] - cat_score_total)
            sd = math.sqrt(var)
            score_result['categories'].append({cat_name: {'cat_question_num': len(cat_questions),
                                                          'cat_score_total': cat_score_total,
                                                          'cat_score_avg': mean,
                                                          'cat_score_sd': sd}})
    # Overall
    sum_question = 0.0
    for item in questions:
        sum_question = item['score'] + sum_question
    mean = sum_question / len(questions)
    single_score_list = []
    for single_score_num in scores_num_list:
        score = 0.0
        for index, value in enumerate(single_score_num):
            score = score + float(value)
        single_score_list.append(score)
    var = 0.0
    for score in single_score_list:
        if questionnaires_name in ['DTDD', 'Empathy']:
            var = var + (score / len(questions) - mean) * (score / len(questions) - mean)
        else:
            var = var + (score - sum_question) * (score - sum_question)
    sd = math.sqrt(var)
    score_result['Overall'] = {'question_num': len(questions),
                               'score_total': sum_question,
                               'score_avg': mean,
                               'score_sd': sd}
    
    # save
    score_result['questions'] = questions
    with open(
            f"results/{model_name}/{questionnaires_name}/score_result_{language}_{run_count}.json",
            'w', encoding='utf') as f:
        if language == 'en':
            json.dump(score_result, f, indent=4)
        elif language == 'zh':
            f.write(json.dumps(score_result, ensure_ascii=False, indent=4))


def result_display(questionnaires_name, model_name, language, run_count):
    """
    Dispalys the result
    @param run_count:
    @param questionnaires_name:
    @param model_name:
    @param language:
    @return:
    """
    # load xlsx
    if os.path.exists(f'results/Dark_Negative_score_{run_count}.xlsx'):
        psy_score = openpyxl.load_workbook(f'results/Dark_Negative_score_{run_count}.xlsx')
    else:
        psy_score = openpyxl.Workbook()
    # get sheets
    sheet = psy_score.create_sheet(questionnaires_name, 0)
    psy_score.active = sheet
    cell_model_name = sheet.cell(1, 2)
    cell_model_name.value = model_name
    cell_crowd_name = sheet.cell(1, 3)
    cell_crowd_name.value = 'crowd'
    with open(f"data/Crowd/{questionnaires_name}/score_result.json", 'r') as crowd_json_file:
        crowd_json_object = json.load(crowd_json_file)
    with open(
            f"results/{model_name}/{questionnaires_name}/score_result_{language}_{run_count}.json",
            'r', encoding='utf-8') as score_result_file:
        result_data = json.load(score_result_file)
    categories_list = result_data['categories']
    categories_list_crowd = crowd_json_object['categories']
    categories_name_list = []
    categories_score_list_avg = []
    categories_score_list_avg_crowd = []
    categories_score_list_sum = []
    categories_score_list_sum_crowd = []
    for category_dicts in categories_list:
        for kk, vv in category_dicts.items():
            categories_name_list.append(kk)
            categories_score_list_avg.append(
                str(round(vv['cat_score_avg'], 1)) + '±' + str(round(vv['cat_score_sd'], 1)))
            categories_score_list_sum.append(
                str(round(vv['cat_score_total'], 1)) + '±' + str(round(vv['cat_score_sd'], 1)))
    for category_dicts in categories_list_crowd:
        for kk, vv in category_dicts.items():
            categories_score_list_avg_crowd.append(
                str(round(vv['cat_score_avg'], 1)) + '±' + str(round(vv['cat_score_sd'], 1)))
            categories_score_list_sum_crowd.append(
                str(round(vv['cat_score_total'], 1)) + '±' + str(round(vv['cat_score_sd'], 1)))
    for i in range(len(categories_name_list)):
        cell = sheet.cell(i + 2, 1)
        cell.value = categories_name_list[i]
    if len(categories_name_list) > 0:
        cell = sheet.cell(len(categories_name_list) + 2, 1)
        cell.value = 'Overall'
        if questionnaires_name in ['DTDD', 'Empathy']:
            overall_score_list_avg_sd = str(round(result_data['Overall']["score_avg"], 1)) + '±' + str(
                round(result_data['Overall']["score_sd"], 1))
            overall_score_list_avg_sd_crowd = str(round(crowd_json_object['Overall']["score_avg"], 1)) + '±' + str(
                round(crowd_json_object['Overall']["score_sd"], 1))
        else:
            overall_score_list_avg_sd = str(round(result_data['Overall']["score_total"], 1)) + '±' + str(
                round(result_data['Overall']["score_sd"], 1))
            overall_score_list_avg_sd_crowd = str(round(crowd_json_object['Overall']["score_total"], 1)) + '±' + str(
                round(crowd_json_object['Overall']["score_sd"], 1))
        cell_model = sheet.cell(len(categories_name_list) + 2, 2)
        cell_crowd = sheet.cell(len(categories_name_list) + 2, 3)
        cell_model.value = overall_score_list_avg_sd
        cell_crowd.value = overall_score_list_avg_sd_crowd
    for i in range(len(categories_score_list_sum)):
        cell_model = sheet.cell(i + 2, 2)
        cell_crowd = sheet.cell(i + 2, 3)
        if questionnaires_name in ['DTDD', 'Empathy']:
            cell_model.value = categories_score_list_avg[i]
            cell_crowd.value = categories_score_list_avg_crowd[i]
        else:
            cell_model.value = categories_score_list_sum[i]
            cell_crowd.value = categories_score_list_sum_crowd[i]
    # Modify style
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['A'].width = 22
    for row in sheet['C']:
        row.font = Font(name='Times New Roman', size=11, color='FF0000')
    for row in sheet['A']:
        row.font = Font(name='Times New Roman', size=11, bold=True)
    for row in sheet[1]:
        row.font = Font(name='Times New Roman', size=11, bold=True)
    psy_score.save(f"results/Dark_Negative_score_{run_count}.xlsx")


def get_access_token(key):
    """
    Generating an Access Token with AK, SK for Baidu AI Cloud
    @return: access_token，or None(when error)
    """
    if key.find(',') != -1 and len(key.split(',')) == 2:
        api_key = key.split(',')[0]
        secret_key = key.split(',')[1]
        url = "https://aip.baidubce.com/oauth/2.0/token"
        params = {"grant_type": "client_credentials", "client_id": api_key, "client_secret": secret_key}
        return str(requests.post(url, params=params).json().get("access_token"))
    else:
        return "Invalid key"


def get_common_prompts_json():
    """
    get common prompts json object
    @return:
    """
    with open(
            'data/common_prompts.json', 'r',
            encoding='utf-8') as common_prompts_json_file:
        common_prompts_json_object = json.load(common_prompts_json_file)
        return common_prompts_json_object
